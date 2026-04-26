import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


def generate_excel_receipt(order):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Чек заказа №{order.pk}"

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.merge_cells('A1:D1')
    ws['A1'] = f'ЧЕК ЗАКАЗА №{order.pk}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A3'] = 'Дата заказа:'
    ws['B3'] = order.created_at.strftime('%d.%m.%Y %H:%M')
    ws['A4'] = 'Покупатель:'
    ws['B4'] = order.user.username
    ws['A5'] = 'Email:'
    ws['B5'] = order.email
    ws['A6'] = 'Телефон:'
    ws['B6'] = order.phone
    ws['A7'] = 'Адрес доставки:'
    ws['B7'] = order.address

    headers = ['Товар', 'Цена за шт.', 'Количество', 'Стоимость']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    row = 10
    for item in order.items.all():
        ws.cell(row=row, column=1, value=item.product.name).border = thin_border
        ws.cell(row=row, column=2, value=float(item.price)).border = thin_border
        ws.cell(row=row, column=3, value=item.quantity).border = thin_border
        ws.cell(row=row, column=4, value=float(item.item_total())).border = thin_border
        row += 1

    ws.cell(row=row, column=3, value='ИТОГО:').font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(order.total_price)).font = Font(bold=True)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output